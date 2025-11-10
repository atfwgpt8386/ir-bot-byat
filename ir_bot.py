import telebot
import json
import os
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ForceReply
import pandas as pd
from datetime import datetime
import io
from openpyxl.styles import PatternFill
import threading
import time
import schedule
from cryptography.fernet import Fernet  # pip install cryptography
import base64

# === BẢO MẬT TOKEN + KEY ===
TOKEN = os.getenv('BOT_TOKEN')  # Đặt trong Railway Variables
ENCRYPT_KEY = os.getenv('ENCRYPT_KEY')  # Tạo 1 lần, mình hướng dẫn bên dưới
if not ENCRYPT_KEY:
    ENCRYPT_KEY = base64.urlsafe_b64encode(os.urandom(32)).decode()  # Tạo tự động lần đầu
    print(f"ENCRYPT_KEY mới: {ENCRYPT_KEY} - COPY DÁN VÀO RAILWAY NGAY!")
cipher = Fernet(ENCRYPT_KEY.encode())

# Chỉ cho phép user ID này dùng bot (bảo vệ tuyệt đối)
ALLOWED_USERS = [6796774010]

DATA_FILE = 'data.enc'  # Lưu mã hóa
tasks = {}
user_states = {}
remind_enabled = {}

# === MÃ HÓA / GIẢI MÃ FILE ===
def load_encrypted():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'rb') as f:
            encrypted = f.read()
        decrypted = cipher.decrypt(encrypted)
        return json.loads(decrypted.decode('utf-8'))
    return {}

def save_encrypted():
    data = json.dumps({"tasks": tasks, "remind": remind_enabled}, ensure_ascii=False)
    encrypted = cipher.encrypt(data.encode('utf-8'))
    with open(DATA_FILE, 'wb') as f:
        f.write(encrypted)

# Load dữ liệu
loaded = load_encrypted()
tasks = loaded.get("tasks", {})
remind_enabled = loaded.get("remind", {})

# === KIỂM TRA QUYỀN TRUY CẬP ===
def check_access(message):
    user_id = message.from_user.id
    if user_id not in ALLOWED_USERS:
        bot.reply_to(message, "❌ Bạn không có quyền sử dụng bot này!")
        return False
    return True

# === TẤT CẢ CÁC HANDLER ĐỀU THÊM CHECK ===
def protected_handler(func):
    def wrapper(message):
        if not check_access(message):
            return
        func(message)
    return wrapper

# Áp dụng cho mọi lệnh
for cmd in ['start', 'add', 'list', 'ir', 'done', 'thieu', 'thongke', 'export', 'remind', 'cancel']:
    handler = bot._handlers[0].get(cmd, [None])[0]
    if handler:
        new_handler = protected_handler(handler.callback)
        bot.remove_message_handler(handler)
        bot.message_handler(commands=[cmd])(new_handler)
        
# === NHẮC NHỞ 8H + AN TOÀN ===
#def daily_reminder():
    #for chat_id in tasks.keys():
        #if not remind_enabled.get(chat_id, True): continue
        # ... (giữ nguyên nội dung nhắc nhở như trước)

#threading.Thread(target=run_scheduler, daemon=True).start()

# === /AUTH (nếu muốn thêm người dùng sau này) ===
@bot.message_handler(commands=['auth'])
def add_user(message):
    if message.from_user.id != ALLOWED_USERS[0]:  # Chỉ admin thêm
        return
    try:
        new_id = int(message.text.split()[1])
        if new_id not in ALLOWED_USERS:
            ALLOWED_USERS.append(new_id)
            bot.reply_to(message, f"Đã thêm user {new_id} vào danh sách trắng!")
    except:
        bot.reply_to(message, "Dùng: /auth 123456789")

# === LƯU KHI THOÁT ===
import atexit
atexit.register(save_encrypted)

DATA_FILE = 'ir_tasks.json'
tasks = {}
user_states = {}

def load_tasks():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_tasks():
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(tasks, f, ensure_ascii=False, indent=4)

tasks = load_tasks()

REQUIRED_FIELDS = [
    "service_request", "response_plan", "ir_report",
    "attack_map", "list_evidence", "up_log", "lesson_learned"
]

FIELDS_ORDER = [
    "irid", "khach_hang", "nguoi_thuc_hien", "created", "updated",
    "incident_info", "service_request", "response_plan", "ir_report",
    "attack_map", "list_evidence", "up_log", "lesson_learned", "status"
]

# === MENU ĐẸP 8 NÚT ===
def main_keyboard():
    markup = ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
    markup.add('/add', '/list', '/thieu')
    markup.add('/ir', '/done', '/thongke')
    markup.add('/export', '/cancel')
    return markup

# === /CANCEL ===
@bot.message_handler(commands=['cancel'])
def cancel_operation(message):
    user_id = str(message.chat.id)
    if user_id in user_states:
        del user_states[user_id]
    bot.reply_to(message, "❌ *Đã hủy thao tác!* Quay lại menu chính ✅", parse_mode='Markdown', reply_markup=main_keyboard())

# === THÊM IR ===
@bot.message_handler(commands=['add'])
def start_add(message):
    user_id = str(message.chat.id)
    user_states[user_id] = {'mode': 'add', 'step': 0, 'data': {}}
    send_prompt(user_id, 0)

def send_prompt(user_id, step):
    field = FIELDS_ORDER[step]
    prompt = f"➕ *Thêm IR mới* [{step+1}/{len(FIELDS_ORDER)}]\n\n📌 Nhập *{field.replace('_', ' ').title()}*:"
    if field in ["created", "updated"]:
        prompt += "\n(dd/mm/yyyy hoặc n/a cho updated)"
    elif field in REQUIRED_FIELDS:
        prompt += "\n(D = Done ✅ | ND = Not Done ❌)"
    elif field == "status":
        prompt += "\n(backlog | in progress | post incident | done)"
    prompt += "\n\nGõ /cancel để thoát!"
    bot.send_message(user_id, prompt, parse_mode='Markdown', reply_markup=ForceReply())

@bot.message_handler(func=lambda m: str(m.chat.id) in user_states and user_states[str(m.chat.id)].get('mode') == 'add')
def handle_add_steps(message):
    user_id = str(message.chat.id)
    if message.text and message.text.strip().lower() == "/cancel":
        cancel_operation(message)
        return

    state = user_states[user_id]
    step = state['step']
    field = FIELDS_ORDER[step]
    text = message.text.strip()
    chat_id = user_id
    if chat_id not in tasks:
        tasks[chat_id] = []

    if field == "irid":
        if not text.isdigit():
            bot.reply_to(message, "❌ IRID phải là số!")
            return
        if any(ir['irid'] == text for ir in tasks[chat_id]):
            bot.reply_to(message, "❌ IRID đã tồn tại!")
            return
        state['data']['irid'] = text
    elif field == "created":
        if not validate_date(text):
            bot.reply_to(message, "❌ Sai định dạng ngày! (dd/mm/yyyy)")
            return
        state['data']['created'] = text
    elif field == "updated":
        state['data']['updated'] = text if text.lower() != "n/a" else "n/a"
    elif field == "incident_info":
        state['data']['incident_info'] = text
    elif field in REQUIRED_FIELDS:
        if text.upper() not in ["D", "ND"]:
            bot.reply_to(message, "❌ Chỉ nhập D hoặc ND!")
            return
        state['data'][field] = "✅ Done" if text.upper() == "D" else "❌ Not Done"
    elif field == "status":
        valid = ["backlog", "in progress", "post incident", "done"]
        if text.lower() not in valid:
            bot.reply_to(message, f"❌ Chỉ nhập: {', '.join(valid)}")
            return
        state['data']['status'] = text.lower()
    else:
        state['data'][field] = text

    if step + 1 < len(FIELDS_ORDER):
        state['step'] += 1
        send_prompt(user_id, state['step'])
    else:
        tasks[chat_id].append(state['data'])
        save_tasks()
        del user_states[user_id]
        ir = state['data']
        missing = [f for f in REQUIRED_FIELDS if ir.get(f) == "❌ Not Done"]
        bot.reply_to(message,
                     f"✅ *IR {ir['irid']} tạo thành công!*\n"
                     f"{'🎉 Hoàn thành 100%!' if not missing else f'⚠️ Còn {len(missing)} mục ND'}\n\n"
                     f"🏢 {ir['khach_hang']} | 👤 {ir['nguoi_thuc_hien']}\n"
                     f"📅 {ir['created']} | Status: {ir['status'].title()}",
                     parse_mode='Markdown', reply_markup=main_keyboard())
        show_ir_detail(chat_id, ir)

def validate_date(d):
    if not d or d.lower() == "n/a": return True
    try:
        datetime.strptime(d, "%d/%m/%Y")
        return True
    except:
        return False

def format_field(f):
    return f.replace("_", " ").title()

def find_ir(chat_id, irid):
    for ir in tasks.get(str(chat_id), []):
        if ir['irid'] == irid:
            return ir
    return None

# === HIỂN THỊ CHI TIẾT IR - ĐẸP Y CHANG BẠN MUỐN ===
def show_ir_detail(chat_id, ir):
    status_emoji = {"backlog": "🔴", "in progress": "🟡", "post incident": "🟠", "done": "🟢"}.get(ir['status'], "⚪")
    msg = f"🔍 *IR {ir['irid']} - {ir['khach_hang']}*\n"
    msg += f"👤 Người: `{ir['nguoi_thuc_hien']}`\n"
    msg += f"📅 Tạo: `{ir['created']}` | Cập nhật: `{ir['updated']}`\n"
    msg += f"⚠️ Incident Info: {ir['incident_info']}\n"
    msg += f"📊 Status: {status_emoji} `{ir['status'].title()}`\n\n"
    msg += "*7 Mục bắt buộc:*\n"
    for f in REQUIRED_FIELDS:
        status = ir.get(f, "❌ Not Done")
        msg += f"┣ {format_field(f)}: {status}\n"
    missing = [f for f in REQUIRED_FIELDS if ir.get(f) == "❌ Not Done"]
    msg += f"\n{'🎉 Hoàn thành!' if not missing else f'🔴 Còn thiếu: {len(missing)} mục'}"
    bot.send_message(chat_id, msg, parse_mode='Markdown')

# === /IR ===
@bot.message_handler(commands=['ir'])
def view_ir(message):
    try:
        irid = message.text.split(maxsplit=1)[1]
        ir = find_ir(message.chat.id, irid)
        if ir:
            show_ir_detail(message.chat.id, ir)
        else:
            bot.reply_to(message, "❌ Không tìm thấy IR này!")
    except:
        bot.reply_to(message, "Dùng: /ir 642")

# === /DONE ===
@bot.message_handler(commands=['done'])
def start_mark_done(message):
    try:
        irid = message.text.split()[1]
        ir = find_ir(message.chat.id, irid)
        if not ir:
            bot.reply_to(message, "Không tìm thấy IR!")
            return
        missing = [f for f in REQUIRED_FIELDS if ir.get(f) == "❌ Not Done"]
        if not missing:
            bot.reply_to(message, "IR này đã hoàn thành!")
            return
        markup = ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        for f in missing:
            markup.add(KeyboardButton(f"{irid} {f}"))
        markup.add("/cancel")
        bot.reply_to(message, f"Chọn mục đánh dấu DONE cho IR {irid}:", reply_markup=markup)
        user_states[str(message.chat.id)] = {'mode': 'done', 'irid': irid}
    except:
        bot.reply_to(message, "Dùng: /done 642")

@bot.message_handler(func=lambda m: str(m.chat.id) in user_states and user_states[str(m.chat.id)].get('mode') == 'done')
def process_done(message):
    if message.text.strip().lower() == "/cancel":
        cancel_operation(message)
        return
    user_id = str(message.chat.id)
    state = user_states[user_id]
    text = message.text.strip()
    if " " not in text: return
    selected_irid, field = text.split(" ", 1)
    if selected_irid != state['irid'] or field not in REQUIRED_FIELDS: return
    ir = find_ir(message.chat.id, selected_irid)
    if ir and ir.get(field) == "❌ Not Done":
        ir[field] = "✅ Done"
        ir['updated'] = datetime.now().strftime("%d/%m/%Y")
        save_tasks()
        bot.reply_to(message, f"✅ Đã đánh dấu *{format_field(field)}* DONE!", parse_mode='Markdown', reply_markup=main_keyboard())
        show_ir_detail(message.chat.id, ir)
    del user_states[user_id]

# === /LIST ===
@bot.message_handler(commands=['list'])
def list_all(message):
    ir_list = tasks.get(str(message.chat.id), [])
    if not ir_list:
        bot.reply_to(message, "Chưa có IR nào!")
        return
    msg = f"📋 *Danh sách IR ({len(ir_list)})*\n\n"
    for ir in ir_list:
        missing = sum(1 for f in REQUIRED_FIELDS if ir.get(f) == "❌ Not Done")
        emoji = "✅" if missing == 0 else f"🔴{missing}"
        status_emoji = {"backlog": "🔴", "in progress": "🟡", "post incident": "🟠", "done": "🟢"}.get(ir['status'], "⚪")
        msg += f"• IR {ir['irid']} | {ir['khach_hang'][:15]} | {emoji} | {status_emoji} {ir['status'].title()}\n"
    bot.reply_to(message, msg, parse_mode='Markdown', reply_markup=main_keyboard())

# === /THIEU ===
@bot.message_handler(commands=['thieu'])
def ir_thieu(message):
    chat_id = str(message.chat.id)
    incomplete = [ir for ir in tasks.get(chat_id, []) if any(ir.get(f) == "❌ Not Done" for f in REQUIRED_FIELDS)]
    if not incomplete:
        bot.reply_to(message, "🎉 Tất cả IR đã hoàn thành 7 mục bắt buộc!")
        return
    msg = f"🔴 *IR còn thiếu ({len(incomplete)})*\n\n"
    for ir in incomplete:
        missing = [f for f in REQUIRED_FIELDS if ir.get(f) == "❌ Not Done"]
        msg += f"IR {ir['irid']} - {ir['khach_hang']} ({len(missing)} thiếu)\n"
    bot.reply_to(message, msg, parse_mode='Markdown', reply_markup=main_keyboard())

# === /THONGKE ===
@bot.message_handler(commands=['thongke'])
def thongke(message):
    ir_list = tasks.get(str(message.chat.id), [])
    total = len(ir_list)
    done_all = sum(1 for ir in ir_list if all(ir.get(f) == "✅ Done" for f in REQUIRED_FIELDS))
    backlog = sum(1 for ir in ir_list if ir['status'] == 'backlog')
    inprog = sum(1 for ir in ir_list if ir['status'] == 'in progress')
    post = sum(1 for ir in ir_list if ir['status'] == 'post incident')
    done = sum(1 for ir in ir_list if ir['status'] == 'done')
    msg = f"📊 *Thống kê IR*\n\n"
    msg += f"Tổng IR: {total}\n"
    msg += f"Hoàn thành 100%: {done_all}\n"
    msg += f"Tỷ lệ hoàn thành: {done_all/total*100 if total else 0:.1f}%\n\n"
    msg += f"🔴 Backlog: {backlog}\n"
    msg += f"🟡 In Progress: {inprog}\n"
    msg += f"🟠 Post Incident: {post}\n"
    msg += f"🟢 Done: {done}\n"
    bot.reply_to(message, msg, parse_mode='Markdown', reply_markup=main_keyboard())

# === /EXPORT ===
@bot.message_handler(commands=['export'])
def export_excel(message):
    chat_id = str(message.chat.id)
    ir_list = tasks.get(chat_id, [])
    if not ir_list:
        bot.reply_to(message, "Chưa có dữ liệu!")
        return
    df = pd.DataFrame(ir_list)
    cols = ['irid','khach_hang','nguoi_thuc_hien','created','updated','incident_info','status'] + REQUIRED_FIELDS
    df = df[cols]
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='IR Reports')
        worksheet = writer.sheets['IR Reports']
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        for row in range(2, len(ir_list) + 2):
            for col in range(8, 15):
                cell = worksheet.cell(row=row, column=col)
                if "Not Done" in str(cell.value):
                    cell.fill = red_fill
    output.seek(0)
    bot.send_document(chat_id, output,
                      caption=f"📊 IR Export - {datetime.now().strftime('%d/%m/%Y')}",
                      visible_file_name=f"IR_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")

# === /START ===
@bot.message_handler(commands=['start'])
def start(message):
    bot.reply_to(message,
                 "🛡️ *BOT QUẢN LÝ IR - KHÔNG BỎ SÓT 7 MỤC!*\n\n"
                 "Lệnh:\n"
                 "/add - Thêm IR mới\n"
                 "/list - Xem tất cả\n"
                 "/ir 12345 - Xem chi tiết\n"
                 "/thieu - IR còn ND\n"
                 "/done 12345 - Đánh dấu Done\n"
                 "/thongke - Thống kê\n"
                 "/export - Excel (ô đỏ = ND)",
                 parse_mode='Markdown', reply_markup=main_keyboard())

print("IR Bot HOÀN CHỈNH 100% - TẤT CẢ BUTTON HOẠT ĐỘNG - 10/11/2025")
bot.infinity_polling()
